#
# (0) All imports for genetic optimization
#

# Standard library imports
import datetime
import json
import logging
import os
import random
import shutil
import time
from abc import ABC, abstractmethod

# Third-party library imports
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import tensorflow
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# External utility modules
from plot_utils import apply_plot_styling

# Ensure necessary packages are installed
try:
    import sklearn
except ImportError:
    os.system("pip install scikit-learn")
    import sklearn

try:
    import optree
except ImportError:
    os.system("pip install optree")
    import optree

# Initialize logging configuration
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

#
# (1) Abstract class for a Genetic Algo
#
class EvolutionaryAlgorithm(ABC):
    """
    Defines key instance variables and abstract instance methods as needed to perform an EA for some optimization problem.
    """

    def __init__(self,
                 population_size=25,
                 mutation_rate=0.02,
                 crossover_rate=0.8,
                 exceptionalism_rate=0.05,
                 **kwargs
                 ):

        self.population_size = population_size
        self.mutation_rate = mutation_rate
        self.crossover_rate = crossover_rate
        self.exceptionalism_rate = exceptionalism_rate
        self.population = self.initialize_population()
        self.fitness = []
        self.num_generations=None

    @abstractmethod
    def initialize_population(self) -> list:
        """
        Must return a list of dictionaries, each dictionary represents an individual.
        """
        pass

    #
    @abstractmethod
    def _compute_fitness(self):
        """
        Compute the fitness of all individuals in the current population.
        This could store the results in an internal data structure or as an attribute of the individuals.
        """
        pass
    #
    @abstractmethod
    def select_parents(self) -> tuple:
        """
        Select two parents for reproduction and return them as a tuple.
        """
        pass

    #
    @abstractmethod
    def crossover(self, parent1: dict, parent2: dict) -> tuple:
        """
        Perform crossover and return two offspring as a tuple of dictionaries.
        """
        pass

    #
    @abstractmethod
    def mutate(self, offspring: dict) -> dict:
        """
        Apply mutation to an offspring and return the mutated individual.
        """
        pass

    #
    def replace_population(self, new_population: list):
      self.population = new_population

    #
    def evolve(self):
        while not self._termination_criterion():

            # Compute fitness of all individuals
            self._compute_fitness()

            new_population = []

            while len(new_population) < len(self.population):
                parent1, parent2 = self.select_parents()

                if random.random() < self.crossover_rate:
                    offspring1, offspring2 = self.crossover(parent1, parent2)
                else:
                    offspring1, offspring2 = parent1, parent2

                if random.random() < self.mutation_rate:
                    offspring1 = self.mutate(offspring1)
                    offspring2 = self.mutate(offspring2)

                new_population.extend([offspring1, offspring2])

            self.replace_population(new_population[:len(self.population)])
    #
    @abstractmethod
    def _termination_criterion(self) -> bool:
        """
        Determines whether the evolutionary algorithm should terminate.
        Return True if termination criteria are met, False otherwise.
        """
        pass

#
# (2) Core Logic for Genetic Opti
#
class VanillaModelConfOptimizer(EvolutionaryAlgorithm):

    def __init__(self, feasible_params, data_assembler=None, train_X=None, train_Y=None, test_X=None, test_Y=None, **kwargs):
        self.forecast_parameters = self._init_forecast_parameters(data_assembler)
        self.feasible_space = self._init_feasible_space(feasible_params)
        self.feasible_space_types = self._determine_param_types(self.feasible_space)
        self.train_X, self.train_Y, self.test_X, self.test_Y = self._init_data(data_assembler, train_X, train_Y, test_X, test_Y)
        self.data_assembler = data_assembler
        self.best_individual = None
        self.max_fitness = float('-inf')
        self.fitness_cache = {}
        self.evaluation_counter = 0
        self.total_individuals_to_evaluate=None
        super().__init__(**kwargs)
        logger.info("[INFO] Initialization complete.")

    def _init_forecast_parameters(self, data_assembler):
        if hasattr(data_assembler, 'get_forecast_parameters'):
            forecast_parameters = data_assembler.get_forecast_parameters()
            logger.debug("[DEBUG] Forecast parameters obtained from data assembler.")
        else:
            forecast_parameters = None
            logger.warning("[WARN] 'get_forecast_parameters' method is not available in 'data_assembler'. Defaulting to None.")
        return forecast_parameters

    def _init_feasible_space(self, feasible_params):
        feasible_space = {}
        for key, values in feasible_params.items():
            if isinstance(values[0], (int, float)):
                feasible_space[key] = (min(values), max(values))
            else:
                feasible_space[key] = values
        logger.debug("[DEBUG] Feasible space initialized.")
        return feasible_space

    def _determine_param_types(self, feasible_space):
        param_types = {}
        for key, space in feasible_space.items():
            if isinstance(space, tuple) and len(space) == 2:
                param_types[key] = "continuous"
            elif isinstance(space, list) or all(isinstance(x, str) for x in space):
                param_types[key] = "discrete"
            else:
                raise ValueError(f"Invalid space for parameter '{key}': {space}")
        logger.debug("[DEBUG] Parameter types determined.")
        return param_types

    def _init_data(self, data_assembler, train_X=None, train_Y=None, test_X=None, test_Y=None):
        self.data_assembler = data_assembler
        if self.data_assembler is not None:
            if hasattr(self.data_assembler, 'get_train_data') and hasattr(self.data_assembler, 'get_test_data'):
                a, b = self.data_assembler.get_train_data()
                c, d = self.data_assembler.get_test_data()
                logger.debug("[DEBUG] Train and test data obtained from data assembler.")
                return a, b, c, d
            elif hasattr(self.data_assembler, 'get_samples'):
                a, b, c, d = self.data_assembler.get_samples()
                logger.debug("[DEBUG] Samples obtained from data assembler.")
                return a, b, c, d
            else:
                logger.error("[ERROR] Data assembler does not provide necessary methods (get_train_data, get_test_data, or get_samples).")
                raise ValueError("Data assembler does not provide necessary methods (get_train_data, get_test_data, or get_samples).")
        elif train_X is not None and train_Y is not None and test_X is not None and test_Y is not None:
            logger.debug("[DEBUG] Data initialized from provided data samples.")
            return train_X, train_Y, test_X, test_Y
        else:
            logger.error("[ERROR] Insufficient data sources provided.")
            raise ValueError("Insufficient data sources provided.")

    def initialize_population(self):
        population = []
        for _ in range(self.population_size):
            individual = {}
            for key, space in self.feasible_space.items():
                param_type = self.feasible_space_types[key]
                if param_type == "continuous":
                    individual[key] = round(random.uniform(space[0], space[1]), 3)
                elif param_type == "discrete":
                    individual[key] = random.choice(space)
            population.append(individual)
        logger.debug("[DEBUG] Population initialized.")
        return population

    def _compute_fitness(self):
        logger.info("[INFO] Computing fitness...")
        fitness_and_rmse_and_time = [self._individual_fitness(individual) for individual in self.population]
        self.fitness_values, self.rmse_values, self.time_values = zip(*fitness_and_rmse_and_time)

        min_fitness = min(value for value, _, _ in fitness_and_rmse_and_time if not np.isnan(value))
        self.fitness_values = [value if not np.isnan(value) else min_fitness for value in self.fitness_values]

        if min_fitness < 0:
            shift = abs(min_fitness)
        else:
            shift = 0

        adjusted_fitness = [value + shift for value in self.fitness_values]
        total_fitness = sum(adjusted_fitness)
        self.relative_fitness = [fitness / total_fitness for fitness in adjusted_fitness]
        logger.info("[INFO] Fitness computed and adjusted.")

    def _control_mechanism(self):
        logger.debug("[DEBUG] Control mechanism check - currently a placeholder.")
        pass

    def _individual_fitness(self, individual):
        self.evaluation_counter += 1
        self._control_mechanism()
        cache_key = self._create_cache_key(individual)
        if cache_key in self.fitness_cache:
            return self.fitness_cache[cache_key]

        start_time = time.time()
        predictions = self.fit_predict(individual)
        elapsed_time = time.time() - start_time

        r2 = self._compute_r2(predictions)
        rmse = self._compute_rmse(predictions)

        self.fitness_cache[cache_key] = (round(r2, 3), round(rmse, 3), round(elapsed_time, 3))

        if r2 > self.max_fitness:
            self._handle_rediscovery_of_best_individual(individual, r2, predictions)

        # Report progress
        progress = self.evaluation_counter / self.total_individuals_to_evaluate
        logger.info(f"[INFO] Evaluation {self.evaluation_counter}/{self.total_individuals_to_evaluate} completed. Progress: {progress:.2%}")

        self._handle_individual_evaluation(individual, r2, rmse, elapsed_time)

        return r2, rmse, elapsed_time

    def _handle_rediscovery_of_best_individual(self, individual, fitness_value, predictions):
        if fitness_value > self.max_fitness:
            self.max_fitness = fitness_value
            self.best_individual = individual
            self.best_predictions = predictions
            logger.info(f"[INFO] New best individual discovered with fitness: {fitness_value:.3f}")

    def _handle_individual_evaluation(self, individual, r2, rmse, elapsed_time):
        logger.debug(f"[DEBUG] Individual evaluation handled. R2: {r2:.3f}, RMSE: {rmse:.3f}, Time: {elapsed_time:.3f} seconds")
        pass

    def _create_cache_key(self, individual):
        return json.dumps(individual, sort_keys=True)

    def _compute_r2(self, predictions):
        ss_res = np.sum((predictions - self.test_Y) ** 2)
        ss_tot = np.sum((self.test_Y - np.mean(self.test_Y)) ** 2)
        if ss_tot == 0:
            logger.error("[ERROR] Division by zero in R2 calculation.")
            raise ValueError("Division by zero in R2 calculation.")
        return 1 - (ss_res / ss_tot)

    def _compute_rmse(self, predictions):
        mse = np.mean((predictions - self.test_Y) ** 2)
        return np.sqrt(mse)

    def select_parents(self):
        parent1 = np.random.choice(self.population, p=self.relative_fitness)
        parent2 = np.random.choice(self.population, p=self.relative_fitness)
        logger.debug("[DEBUG] Parents selected for crossover.")
        return parent1, parent2

    def crossover(self, parent1, parent2):
        offspring1, offspring2 = {}, {}
        for key in parent1:
            if random.random() < 0.5:
                offspring1[key], offspring2[key] = parent1[key], parent2[key]
            else:
                offspring1[key], offspring2[key] = parent2[key], parent1[key]
        logger.debug("[DEBUG] Crossover performed.")
        return offspring1, offspring2

    def mutate(self, offspring):
        gene_to_mutate = random.choice(list(offspring.keys()))
        param_type = self.feasible_space_types[gene_to_mutate]

        if param_type == "continuous":
            min_val, max_val = self.feasible_space[gene_to_mutate]
            offspring[gene_to_mutate] = round(random.uniform(min_val, max_val), 3)
        elif param_type == "discrete":
            offspring[gene_to_mutate] = random.choice(self.feasible_space[gene_to_mutate])
        logger.debug("[DEBUG] Mutation performed.")
        return offspring

    def replace_population(self, new_population):
        self._recompute_exceptional_individuals()
        for idx in self.exceptional_individuals:
            new_population[idx] = self.population[idx]
        self.population = new_population
        logger.debug("[DEBUG] Population replaced.")

    def _recompute_exceptional_individuals(self):
        num_exceptional = int(self.exceptionalism_rate * len(self.population))
        self.exceptional_individuals = np.argsort(self.fitness_values)[-num_exceptional:]
        logger.debug("[DEBUG] Exceptional individuals recomputed.")

    def evolve(self, num_generations):
        self.num_generations=num_generations
        # Set the total number of evaluations based on the population size and number of generations
        self.total_individuals_to_evaluate = self.population_size * num_generations

        self.current_generation = 0
        while not self._termination_criterion(num_generations):
            logger.info(f"[INFO] Starting generation {self.current_generation}")
            self._single_iteration()
            self.current_generation += 1
            self._handle_replacement_of_population()

        logger.info(f"[INFO] Evolution completed. Total evaluations: {self.evaluation_counter}/{self.total_individuals_to_evaluate}")

    def _single_iteration(self):
        self._compute_fitness()
        new_population = []
        while len(new_population) < len(self.population):
            parent1, parent2 = self.select_parents()
            if random.random() < self.crossover_rate:
                offspring1, offspring2 = self.crossover(parent1, parent2)
            else:
                offspring1, offspring2 = parent1, parent2
            if random.random() < self.mutation_rate:
                offspring1 = self.mutate(offspring1)
                offspring2 = self.mutate(offspring2)
            new_population.extend([offspring1, offspring2])
        self.replace_population(new_population[:len(self.population)])

    def _termination_criterion(self, num_generations):
        return self.current_generation >= num_generations

    def _handle_replacement_of_population(self):
        average_fitness = sum(self.fitness_values) / len(self.fitness_values)
        logger.info(f"[INFO] After {self.current_generation} EA iterations, average fitness of the replaced population is {average_fitness:.3f}")

    @abstractmethod
    def fit_predict(self, individual):
        pass

#
# (3) Diagnostic metrics
#
class MonitoredModelConfOptimizer(VanillaModelConfOptimizer):
    """
    Inherits from Vainilla. But adds instance variables to keep track of the trend in parameters of fitness distribution.
    """

    #
    # (0)
    #
    def __init__(self, feasible_params, data_assembler=None, train_X=None, train_Y=None, test_X=None, test_Y=None, **kwargs):
        super().__init__(feasible_params, data_assembler, train_X, train_Y, test_X, test_Y, **kwargs)
        self.percentile_25_values = []
        self.percentile_50_values = []
        self.percentile_75_values = []
        self.solutions = []
    #
    # (1.1)
    #
    def _handle_individual_evaluation(self, individual, r2, rmse, elapsed_time):
        # Store the individual and its performance metrics in the solutions list
        self.solutions.append({
            "params": individual,
            "metrics": {"r2": r2,
                        "rmse": rmse,
                        "elapsed_time":elapsed_time
                        }
        })
        # Display updated plot after each individual evaluation
        self.plot_best_solutions()
    #
    # (1.2)
    #
    def plot_best_solutions(self, max_rank_size=20):
        # Filter out solutions outside the R2 range -1 to 1
        valid_solutions = [solution for solution in self.solutions if -1 <= solution["metrics"]["r2"] <= 1]

        # Sort solutions by R2 value in descending order and limit the number of solutions
        sorted_solutions = sorted(valid_solutions, key=lambda s: s["metrics"]["r2"], reverse=True)[:max_rank_size]

        # Extract R2 values and parameter strings
        r2_values = [solution["metrics"]["r2"] for solution in sorted_solutions]
        param_strings = [
            '<br>'.join([f'{key}: {value}' for key, value in solution["params"].items()])
            for solution in sorted_solutions
        ]

        # Create the bar plot
        fig = go.Figure(data=[
            go.Bar(
                y=list(range(len(sorted_solutions))),
                x=r2_values,
                orientation='h',
                marker=dict(
                    color='blue',  # Uniform color for bars
                    line=dict(color='white', width=2),
                    opacity=0.7
                ),
                hovertext=param_strings,
                hoverinfo='text+x'  # Show custom hover text and x-value
            )
        ])

        # Adjust the layout
        fig.update_layout(
            title="Ranking of Best Solutions",
            xaxis_title="Fitness (R2 Value)",
            yaxis=dict(
                showticklabels=False,  # Omit axis labels for each bar
                autorange="reversed"  # Reverse the y-axis to show the best solution at the top
            )
        )

        fig.show()
        return fig

    #
    # (2.1)
    #
    def _handle_replacement_of_population(self):
        # Update fitness percentiles
        self.percentile_25_values.append(np.percentile(self.fitness_values, 25))
        self.percentile_50_values.append(np.percentile(self.fitness_values, 50))
        self.percentile_75_values.append(np.percentile(self.fitness_values, 75))

        # Print or log information about the replaced population
        average_fitness = sum(self.fitness_values) / len(self.fitness_values)
        print(f"After {self.current_generation} EA iterations, average fitness of the replaced population is {average_fitness:.4f}")

        # Display updated plot after each individual evaluation
        self.plot_fitness()
    #
    # (2.2)
    #
    def plot_fitness(self):
        fig = go.Figure()

        marker_params = {"size": 10, "opacity": 0.7}
        line_params = {"width": 2, "dash": 'dash'}

        fig.add_trace(go.Scatter(x=list(range(len(self.percentile_25_values))),
                                 y=self.percentile_25_values,
                                 mode='lines+markers',
                                 name='25th Percentile',
                                 line=dict(color='blue', **line_params),
                                 marker=marker_params))

        fig.add_trace(go.Scatter(x=list(range(len(self.percentile_50_values))),
                                 y=self.percentile_50_values,
                                 mode='lines+markers',
                                 name='50th Percentile',
                                 line=dict(color='green', **line_params),
                                 marker=marker_params))

        fig.add_trace(go.Scatter(x=list(range(len(self.percentile_75_values))),
                                 y=self.percentile_75_values,
                                 mode='lines+markers',
                                 name='75th Percentile',
                                 line=dict(color='yellow', **line_params),
                                 marker=marker_params))

        fig.update_layout(title="Evolution of Population Fitness",
                          xaxis_title="Generation",
                          yaxis_title="Fitness Value",
                          xaxis=dict(tickvals=list(range(len(self.percentile_25_values))), type="category"),
                          legend=dict(x=0.02, y=0.5, bgcolor='rgba(255,255,255,0.5)', font=dict(size=9)))

        # Apply consistent styling from StructuredTimeSeriesCollection
        fig = apply_plot_styling(fig)
        fig.show()
        return fig
    #
    # (3)
    #
    def _handle_rediscovery_of_best_individual(self, individual, fitness_value, predictions):
        #
        if fitness_value > self.max_fitness:
            self.max_fitness = fitness_value
            self.best_individual = individual
            self.best_predictions = predictions  # Store the best predictions

            print(f"New best individual discovered with fitness: {fitness_value:.4f}")

            # Delegate the plotting of best predictions to data_assembler
            if hasattr(self, 'data_assembler') and callable(getattr(self.data_assembler, 'plot', None)):
                self.data_assembler.plot(self.best_predictions)
            else:
                print("Data assembler does not support plotting or is not available.")

#
# (4) Exporting optimization output
#
class ServerModelConfOptimizer(MonitoredModelConfOptimizer):

    """
    It takes on from the plotting capabilities of UXModelConfOptimizer. It's built to save some reports to the home dir of a local server.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.start_time = datetime.datetime.now()  # Initialize start time
        self.export_results_called = False  # Flag to track if export_results has been called

        # Retrieve results folder from kwargs or use default
        results_folder = kwargs.get('results_folder', None)

        # If no results folder is provided, assume the home directory
        if results_folder:
            self.home_public_results_dir = results_folder
        else:
            self.home_public_results_dir = os.path.join(os.path.expanduser("~"), "EA_forecast_optimization_results")

        self.home_protected_results_dir = os.path.join(self.home_public_results_dir, "please_dont_open")

        # Create directories if they don't exist
        os.makedirs(self.home_public_results_dir, exist_ok=True)
        os.makedirs(self.home_protected_results_dir, exist_ok=True)

        print(f"Results will be saved in: {self.home_public_results_dir}")
    #
    # 1.
    #
    def _handle_rediscovery_of_best_individual(self, individual, fitness_value, predictions):
        super()._handle_rediscovery_of_best_individual(individual, fitness_value, predictions)
        fig = self.data_assembler.plot(self.best_predictions)
        self._export_plot(fig, "best_predictions_plot")
        self.export_results(event="new_best_solution")
    #
    # 2.
    #
    def _handle_replacement_of_population(self):
        super()._handle_replacement_of_population()
        fig =self.plot_fitness()
        self._export_plot(fig, "fitness_evolution_plot")
        self.export_results(event="new_generation")
    #
    def _export_plot(self, fig, plot_name):
        # File path for the plot
        plot_file = os.path.join(self.home_public_results_dir, f"{self.__class__.__name__}_{plot_name}.html")
        # Export the plot as an HTML file
        pio.write_html(fig, file=plot_file, auto_open=False)
    #
    # 3.
    #
    def export_results(self, event):
        #
        # (0) clean previous results the same class
        #
        if not self.export_results_called:
            #
            self._clean_old_files(self.home_public_results_dir)
            self._clean_old_files(self.home_protected_results_dir)
            self.export_results_called = True
            self.baseline_percentiles_fitness=None
            self.baseline_percentiles_rmse=None
        #
        # (1) Mandatory copy on protected dir
        #

        # Use 'please_dont_open' as results_dir for further processing
        file_name = f"{self.__class__.__name__}_results.xlsx"
        file_path = os.path.join(self.home_protected_results_dir, file_name)

        # Handling workbook creation or loading
        if not os.path.exists(file_path):
            # Create new workbook and populate initial sheets if first call
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet
            self._populate_optimization_parameters_sheet(workbook)
            self._populate_search_space_sheet(workbook)
        else:
            # Load existing workbook
            workbook = openpyxl.load_workbook(file_path)

        # Update the corresponding sheet based on the event
        if event == "new_best_solution":
            self._update_best_solution_sheet(workbook)
            self._update_best_forecast_sheet(workbook)
        elif event == "new_generation":
            self._update_general_results_sheet(workbook)
            self._update_general_results_rmse_sheet(workbook)

        # Apply styles to all sheets before saving
        self._apply_styles_to_all_sheets(workbook)

        # Save the workbook in the child directory
        workbook.save(file_path)
        #
        # (2) Non-Mandatory copy on public dir
        #
        try:
            self._make_read_only_copy(file_path, self.home_public_results_dir)
        except Exception as e:
            print(f"Error occurred while copying the file: {e}")
    #
    def _clean_old_files(self, directory):
        for file in os.listdir(directory):
            if file.startswith(self.__class__.__name__):
                os.remove(os.path.join(directory, file))
    #
    def _make_read_only_copy(self, source_path, destination_dir):
        destination_path = os.path.join(destination_dir, os.path.basename(source_path))
        shutil.copyfile(source_path, destination_path)
        #os.chmod(destination_path, 0o444)  # Set file to read-only
    #
    def _format_elapsed_time(self):
        # Calculate and format the elapsed time since instantiation
        elapsed_time = datetime.datetime.now() - self.start_time
        hours, remainder = divmod(elapsed_time.total_seconds(), 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{int(hours)}h {int(minutes)}m {int(seconds)}s"
    #
    def _apply_styles_to_sheet(self, sheet):
        # Style for normal thick (medium) border
        medium_border = Border(left=Side(style='medium', color='000000'),
                            right=Side(style='medium', color='000000'),
                            top=Side(style='medium', color='000000'),
                            bottom=Side(style='medium', color='000000'))

        # Apply auto-adjust columns and medium border
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column if cell.value]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                cell.border = medium_border
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    #
    def _apply_styles_to_all_sheets(self, workbook):
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            self._apply_styles_to_sheet(sheet)

    #
    # (1)
    #
    def _populate_optimization_parameters_sheet(self, workbook):
        sheet = workbook.create_sheet("Optimization Parameters")

        # Add headers
        sheet["A1"] = "Parameter"
        sheet["B1"] = "Value"

        # Define parameters
        params = {
            'Population Size': self.population_size,
            'Number of Generations': self.num_generations,
            'Mutation Rate': self.mutation_rate,
            'Crossover Rate': self.crossover_rate,
            'Start Time': self.start_time.strftime("%Y-%m-%d %H:%M:%S"),
            'Optimizer Class Name': self.__class__.__name__
        }

        # Populate values, starting from row 2
        for i, (key, value) in enumerate(params.items(), start=2):
            sheet[f'A{i}'] = key
            sheet[f'B{i}'] = value
    #
    # (2)
    #
    def _populate_search_space_sheet(self, workbook):
        sheet = workbook.create_sheet("Search Space")

        # Add headers
        sheet["A1"] = "Parameter"
        sheet["B1"] = "Feasible Range"

        # Populate values, starting from row 2
        for i, (param, range_values) in enumerate(self.feasible_space.items(), start=2):
            sheet[f'A{i}'] = param
            sheet[f'B{i}'] = str(range_values)

    #
    # (3)
    #
    def _update_best_solution_sheet(self, workbook):
        if "Best Solution" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Best Solution")
        else:
            sheet = workbook["Best Solution"]

        # Add headers
        sheet["A1"] = "Parameter"
        sheet["B1"] = "Value"

        # Populate the best individual parameters
        for i, (param, value) in enumerate(self.best_individual.items(), start=2):  # Start from row 2
            sheet[f'A{i}'] = param
            sheet[f'B{i}'] = value

        # Add additional details
        sheet[f'A{i+1}'] = 'Fitness'
        sheet[f'B{i+1}'] = self.max_fitness
        sheet[f'A{i+2}'] = 'Elapsed Time'
        sheet[f'B{i+2}'] = self._format_elapsed_time()

    #
    # (4)
    #

    def _update_general_results_sheet(self, workbook, sheet_name="General Results-r2"):
        # Calculating current percentiles and max fitness
        percentiles = {
            "25th": np.percentile(self.fitness_values, 25),
            "50th": np.percentile(self.fitness_values, 50),
            "75th": np.percentile(self.fitness_values, 75),
            "Max": max(self.fitness_values)
        }

        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # Extended header with increment factors
            sheet.append(["Generation", "Elapsed Time", "25th Percentile", "25th Increment", "50th Percentile", "50th Increment", "75th Percentile", "75th Increment", "Max Fitness", "Max Increment"])

            # Check if baselines are set, if not set them (first iteration)
            self.baseline_percentiles_fitness = percentiles

        else:

            sheet = workbook[sheet_name]


        # Prepare data row with current values and increment factors
        data_row = [
            self.current_generation,
            self._format_elapsed_time(),
        ]

        for key in ["25th", "50th", "75th", "Max"]:
            current_value = round(percentiles[key], 5)
            baseline = self.baseline_percentiles_fitness[key]
            increment_factor = round(current_value / baseline, 5) if baseline != 0 else 'N/A'
            data_row.extend([current_value, increment_factor])

        # Append the data row
        sheet.append(data_row)

    def _update_general_results_rmse_sheet(self, workbook, sheet_name="General Results-rmse"):
        # Calculating current percentiles and max fitness
        percentiles = {
            "25th": np.percentile(self.rmse_values, 25),
            "50th": np.percentile(self.rmse_values, 50),
            "75th": np.percentile(self.rmse_values, 75),
            "Min": min(self.rmse_values)
        }

        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            # Extended header with RMSE and increment factors
            sheet.append(["Generation", "Elapsed Time", "25th Percentile", "25th Increment", "50th Percentile", "50th Increment", "75th Percentile", "75th Increment", "Min rmse", "Min Increment"])
            self.baseline_percentiles_rmse = percentiles
        else:
            sheet = workbook[sheet_name]


        # Prepare data row with current values and increment factors
        data_row = [
            self.current_generation,
            self._format_elapsed_time(),
        ]

        for key in ["25th", "50th", "75th", "Min"]:
            current_value = round(percentiles[key], 5)
            baseline = self.baseline_percentiles_rmse[key]
            increment_factor = round(current_value / baseline, 5) if baseline != 0 else 'N/A'
            data_row.extend([current_value, increment_factor])

        # Append the data row
        sheet.append(data_row)

    #
    # (5)
    #
    def _update_best_forecast_sheet(self, workbook):
        # Create or access the 'Best Forecast' sheet
        if "Best Forecast" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Best Forecast")
        else:
            sheet = workbook["Best Forecast"]

        # Clear the sheet before updating
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        # Prepare the DataFrame
        forecast_data = pd.DataFrame({
            "Historical": self.test_Y,
            "Forecasts": self.best_predictions
        })

        # Write DataFrame to the Excel sheet
        for r_idx, row in enumerate(dataframe_to_rows(forecast_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                # Apply normal border to each cell
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )

        # Optionally, apply any additional styling if necessary
        # self._apply_styles_to_sheet(sheet)
